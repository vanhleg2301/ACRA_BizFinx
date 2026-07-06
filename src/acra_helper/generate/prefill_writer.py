"""
Write pre-fill output: Excel workbook (human-readable) + JSON (machine-readable).
Layout mirrors BizFinx data-entry order: element | context | period | value.
"""
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL = PatternFill(fill_type="solid", fgColor="D6E4F0")


def write_prefill(
    facts: list[dict],
    contexts: dict[str, dict],
    output_dir: Path,
    uen: str,
) -> tuple[Path, Path]:
    """
    Args:
        facts:      [{element, context, value}, ...]  (all facts including text)
        contexts:   context_key → {period, dimension}
        output_dir: destination directory
        uen:        company UEN (used in filename)

    Returns: (xlsx_path, json_path)
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / f"{uen}_prefill.xlsx"
    json_path = output_dir / f"{uen}_prefill.json"

    _write_xlsx(facts, contexts, xlsx_path)
    _write_json(facts, json_path)

    return xlsx_path, json_path


def _period_label(ctx_def: dict) -> str:
    p = ctx_def["period"]
    if p["type"] == "instant":
        return p["date"]
    return f"{p['start']} to {p['end']}"


def _dim_label(ctx_def: dict) -> str:
    d = ctx_def.get("dimension")
    if not d:
        return ""
    return f"{d['axis']}={d['member']}"


def _write_xlsx(facts: list[dict], contexts: dict[str, dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pre-fill"

    headers = ["Element", "Context", "Period", "Dimension", "Value"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, fact in enumerate(facts, 2):
        ctx_key = fact.get("context", "")
        ctx_def = contexts.get(ctx_key, {})
        fill = _ALT_FILL if row_idx % 2 == 0 else None

        values: list[Any] = [
            fact.get("element", ""),
            ctx_key,
            _period_label(ctx_def) if ctx_def else ctx_key,
            _dim_label(ctx_def) if ctx_def else "",
            fact.get("value", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fill:
                cell.fill = fill

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 20
    wb.save(path)


def _write_json(facts: list[dict], path: Path) -> None:
    path.write_text(json.dumps(facts, indent=2, ensure_ascii=False), encoding="utf-8")
